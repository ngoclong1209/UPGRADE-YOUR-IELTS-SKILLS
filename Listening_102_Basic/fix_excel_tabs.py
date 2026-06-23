import openpyxl

excel_path = '/Users/vungoclong/Desktop/Antigravity/UPGRADE YOUR ILETS LISTENING/UPGRADE YOUR ILETS LISTENING.xlsx'

wb = openpyxl.load_workbook(excel_path)

# Try to rename back to the exact names user requested if needed, or just use the existing ones
target_order = ['Students', 'Scores', 'DS Bài Học', 'Danh sách link bài học', 'Submissions', 'Kế hoạch', 'Thống kê học viên']

# Keep only the ones that exist, preserving 'Students', 'Scores', etc.
ordered_sheets = []
for name in target_order:
    if name in wb.sheetnames:
        ordered_sheets.append(wb[name])

# Append any remaining sheets (like the HV_ logs) at the end
for name in wb.sheetnames:
    if wb[name] not in ordered_sheets:
        ordered_sheets.append(wb[name])

wb._sheets = ordered_sheets
wb.save(excel_path)
print("Excel tabs reordered!")
