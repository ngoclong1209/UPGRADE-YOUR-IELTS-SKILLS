import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

excel_path = '/Users/vungoclong/Desktop/Antigravity/UPGRADE YOUR ILETS LISTENING/UPGRADE YOUR ILETS LISTENING.xlsx'

# Load workbook
wb = openpyxl.load_workbook(excel_path)

# Ensure sheets exist
target_order = ['Students', 'DS Bài Học', 'Submissions', 'Kế hoạch', 'Thống kê học viên']
# Wait, user specifically said: "DS Bài Học, Submissions, Kế hoạch, Thống kê học viên phải để ngay sau tab Students (mặc định)"
# Let's rename existing tabs if they don't match perfectly, or just map them.
name_mapping = {
    '🔗 Danh sách link bài học': 'DS Bài Học',
    '📅 Kế hoạch 30 ngày': 'Kế hoạch',
    '📊 Thống kê học viên': 'Thống kê học viên'
}
for old_name, new_name in name_mapping.items():
    if old_name in wb.sheetnames:
        wb[old_name].title = new_name

# If "DS Bài Học" doesn't exist, create it
for s_name in target_order:
    if s_name not in wb.sheetnames:
        wb.create_sheet(s_name)

# Reorder
wb._sheets = [wb[x] for x in target_order] + [wb[x] for x in wb.sheetnames if x not in target_order]

# Rebuild 'Kế hoạch'
plan_sheet = wb['Kế hoạch']
plan_sheet.delete_rows(1, plan_sheet.max_row + 1) # Clear existing

headers = ['Ngày Học', 'Tên Bài Tập', 'Link Truy Cập', 'Trạng Thái']
plan_sheet.append(headers)

# Format headers
header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
for col in range(1, 5):
    cell = plan_sheet.cell(row=1, column=col)
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# Data prep
basics = [(f"Basic_Lesson_{i}", f"Basic/Lesson_{i}") for i in range(1, 35)]
inters = [(f"Intermediate_Lesson_{i}", f"Intermediate/Lesson_{i}") for i in range(1, 35)]
advs = [(f"Advanced_Lesson_{i}", f"Advanced/Lesson_{i}") for i in range(1, 35)]

all_lessons = []
for i in range(34):
    all_lessons.append(basics[i])
    all_lessons.append(inters[i])
    all_lessons.append(advs[i])

start_date = datetime(2026, 6, 16)
num_days = 30
lessons_per_day = 102 // num_days
remainder = 102 % num_days

lesson_idx = 0
current_row = 2

# Add Data Validation for Status
dv = DataValidation(type="list", formula1='"☐ Chưa Làm,☑ Hoàn Thành"', allow_blank=True)
plan_sheet.add_data_validation(dv)

for day in range(num_days):
    current_date = start_date + timedelta(days=day)
    date_str = current_date.strftime("%d/%m/%Y")
    
    daily_count = lessons_per_day + (1 if day < remainder else 0)
    for _ in range(daily_count):
        if lesson_idx < len(all_lessons):
            name, folder = all_lessons[lesson_idx]
            url = f"https://ngoclong1209.github.io/pte-listening-audios/{folder}/"
            
            plan_sheet.cell(row=current_row, column=1, value=date_str)
            plan_sheet.cell(row=current_row, column=2, value=name)
            
            link_cell = plan_sheet.cell(row=current_row, column=3, value="Vào Học Ngay")
            link_cell.hyperlink = url
            link_cell.font = Font(color="0563C1", underline="single")
            
            status_cell = plan_sheet.cell(row=current_row, column=4, value="☐ Chưa Làm")
            dv.add(status_cell)
            
            current_row += 1
            lesson_idx += 1

plan_sheet.column_dimensions['A'].width = 15
plan_sheet.column_dimensions['B'].width = 30
plan_sheet.column_dimensions['C'].width = 20
plan_sheet.column_dimensions['D'].width = 20

wb.save(excel_path)
print("Excel reorganized and Kế hoạch rebuilt!")
