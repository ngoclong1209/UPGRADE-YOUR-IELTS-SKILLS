import os
import re
import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink

def get_lessons(base_dir, level):
    folder = os.path.join(base_dir, "ReadingA1-C2", "frontend", "data", level)
    files = [f for f in os.listdir(folder) if f.endswith(".js") and f.startswith("lesson_")]
    # Extract number and sort
    files.sort(key=lambda x: int(re.search(r'lesson_(\d+)', x).group(1)))
    return files

def add_business_days(start_date, add_days):
    current_date = start_date
    while add_days > 0:
        current_date += datetime.timedelta(days=1)
        if current_date.weekday() < 5: # Monday to Friday
            add_days -= 1
    return current_date

def generate_plan():
    print("=" * 50)
    print("🚀 TOOL LÊN LỊCH HỌC READING 9.0 IELTS (4 THÁNG)")
    print("=" * 50)
    
    student_name = input("Nhập tên học viên (phải trùng tên Sheet nếu đã có): ").strip()
    start_date_str = input("Nhập ngày bắt đầu (YYYY-MM-DD): ").strip()
    
    try:
        current_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
    except ValueError:
        print("❌ Lỗi: Sai định dạng ngày. Vui lòng nhập YYYY-MM-DD.")
        return

    # Check if starting on weekend, push to Monday
    if current_date.weekday() >= 5:
        current_date = add_business_days(current_date, 1 if current_date.weekday() == 6 else 2)

    base_dir = os.getcwd()
    
    # 1. Thu thập bài học
    print("\nĐang quét danh sách bài học...")
    levels = {
        "A1-A2": {"files": get_lessons(base_dir, "A1-A2"), "quota": 15},
        "B1-B2": {"files": get_lessons(base_dir, "B1-B2"), "quota": 15},
        "C1-C2": {"files": get_lessons(base_dir, "C1-C2"), "quota": 13}
    }
    
    total_lessons = sum(len(lvl["files"]) for lvl in levels.values())
    print(f"Tổng cộng tìm thấy: {total_lessons} bài.")
    
    # 2. Xây dựng lịch trình
    schedule = []
    
    for level, data in levels.items():
        files = data["files"]
        quota = data["quota"]
        
        chunk = []
        for file in files:
            chunk.append(file)
            if len(chunk) == quota:
                # Flush chunk
                schedule.append((current_date, level, list(chunk)))
                chunk = []
                current_date = add_business_days(current_date, 1)
        
        # Flush remaining
        if chunk:
            schedule.append((current_date, level, list(chunk)))
            current_date = add_business_days(current_date, 1)

    print(f"Lịch học trải dài qua {len(schedule)} ngày học (Không tính T7, CN).")
    print(f"Ngày kết thúc dự kiến: {current_date.strftime('%Y-%m-%d')}")
    
    # 3. Ghi vào Excel
    excel_path = "/Users/vungoclong/Library/CloudStorage/GoogleDrive-ngoclong137129@gmail.com/My Drive/IELTS_LISTENING/UPGRADE YOUR ILETS LISTENING.xlsx"
    print("\nĐang mở file Excel...")
    wb = load_workbook(excel_path)
    
    if student_name in wb.sheetnames:
        ws = wb[student_name]
        print(f"✅ Tìm thấy sheet '{student_name}'. Sẽ ghi nối tiếp dữ liệu...")
    else:
        # Clone template if 'Kế hoạch' exists
        if 'Kế hoạch' in wb.sheetnames:
            ws = wb.copy_worksheet(wb['Kế hoạch'])
            ws.title = student_name
            print(f"✅ Đã tạo sheet mới '{student_name}' từ template 'Kế hoạch'.")
        else:
            ws = wb.create_sheet(student_name)
            ws.append(['Ngày Học', 'Tên Bài Tập', 'Link Truy Cập', 'Trạng Thái'])
            print(f"✅ Đã tạo sheet trống mới '{student_name}'.")

    # Find the last row with data in col A or B
    last_row = ws.max_row
    while last_row > 1 and ws.cell(row=last_row, column=1).value is None and ws.cell(row=last_row, column=2).value is None:
        last_row -= 1

    start_row = last_row + 1
    base_url = "https://ngoclong1209.github.io/ielts-reading-app/frontend/index.html"
    
    print(f"Bắt đầu ghi dữ liệu từ dòng {start_row}...")
    
    for i, (date, level, files) in enumerate(schedule):
        for f in files:
            row_idx = start_row
            ws.cell(row=row_idx, column=1).value = date.strftime("%Y-%m-%d")
            ws.cell(row=row_idx, column=2).value = f"Reading {level} - {f.replace('.js', '')}"
            
            # URL Link
            file_url = f"{base_url}?file={level}/{f}"
            cell_link = ws.cell(row=row_idx, column=3)
            cell_link.value = "🔗 Làm bài ngay"
            cell_link.hyperlink = file_url
            cell_link.style = "Hyperlink"
            
            ws.cell(row=row_idx, column=4).value = "Chưa làm"
            
            start_row += 1

    print("Đang lưu file...")
    wb.save(excel_path)
    print(f"🎉 HOÀN THÀNH! Lịch học đã được thêm vào file {excel_path} cho học viên {student_name}.")

if __name__ == "__main__":
    generate_plan()
