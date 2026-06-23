import os
import re
from openpyxl import load_workbook

def get_lessons(base_dir, level):
    folder = os.path.join(base_dir, "ReadingA1-C2", "frontend", "data", level)
    if not os.path.exists(folder):
        return []
    files = [f for f in os.listdir(folder) if f.endswith(".js") and f.startswith("lesson_")]
    files.sort(key=lambda x: int(re.search(r'lesson_(\d+)', x).group(1)))
    return files

def inject_data():
    base_dir = os.getcwd()
    excel_path = "UPGRADE YOUR ILETS SKILLS.xlsx"
    
    print(f"Opening {excel_path}...")
    wb = load_workbook(excel_path)
    
    sheet_name = "Reading_Data"
    if sheet_name in wb.sheetnames:
        print(f"Sheet {sheet_name} exists, clearing it...")
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        print(f"Creating new sheet {sheet_name}...")
        ws = wb.create_sheet(sheet_name)
        # Put it somewhere near the beginning
        wb.move_sheet(ws, offset=-len(wb.sheetnames)+1)
        
    # Write headers
    ws.append(["Level", "File Name", "Lesson Title"])
    
    levels = ["A1-A2", "B1-B2", "C1-C2"]
    for level in levels:
        files = get_lessons(base_dir, level)
        for f in files:
            title = f.replace('.js', '')
            ws.append([level, f, f"Reading {level} - {title}"])
            
    print("Saving...")
    wb.save(excel_path)
    print("Done! Data injected successfully.")

if __name__ == "__main__":
    inject_data()
