import pandas as pd
from openpyxl import load_workbook
import openpyxl

def normalize_excel():
    excel_path = "UPGRADE YOUR ILETS SKILLS.xlsx"
    print(f"Normalizing {excel_path}...")
    
    wb = load_workbook(excel_path)
    sheet_names = wb.sheetnames
    
    df = pd.read_excel(excel_path, sheet_name="Students")
    
    rename_map = {
        "Mã Học Viên": "Mã số",
        "Tên Học Viên": "Họ và Tên",
        "Tên HV": "Họ và Tên",
        "Mã HV": "Mã số"
    }
    df.rename(columns=rename_map, inplace=True)
    
    new_headers = [
        "Họ và Tên", 
        "Mã số", 
        "Device ID", 
        "Ngày Bắt Đầu", 
        "Reading 1323 (Tick)", 
        "Listening 102 (Tick)", 
        "Full Tests (Tick)", 
        "Tạo Kế Hoạch (Tick)"
    ]
    
    for col in new_headers:
        if col not in df.columns:
            df[col] = ""
            
    df = df[new_headers]
    
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name="Students", index=False)
        
        if "Submissions" not in sheet_names:
            sub_df = pd.DataFrame(columns=["Timestamp", "Student ID", "Lesson", "Score", "Score %"])
            sub_df.to_excel(writer, sheet_name="Submissions", index=False)
            
    # Clean up HV_ sheets if any exist using openpyxl
    wb2 = load_workbook(excel_path)
    for name in wb2.sheetnames:
        if name.startswith("HV_"):
            del wb2[name]
    wb2.save(excel_path)
    
    print("Excel normalized successfully.")

if __name__ == "__main__":
    normalize_excel()
