import os
import re
import pandas as pd
from openpyxl import load_workbook

def get_reading_lessons(base_dir, level):
    folder = os.path.join(base_dir, "ReadingA1-C2", "frontend", "data", level)
    if not os.path.exists(folder):
        return []
    files = [f for f in os.listdir(folder) if f.endswith(".js") and f.startswith("lesson_")]
    files.sort(key=lambda x: int(re.search(r'lesson_(\d+)', x).group(1)))
    return files

def inject_data():
    base_dir = os.getcwd()
    excel_path = "UPGRADE YOUR ILETS SKILLS.xlsx"
    
    print(f"Opening {excel_path}...")
    wb = load_workbook(excel_path)
    
    # 1. Reading_Data
    sheet_name = "Reading_Data"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)
    
    ws.append(["Package", "Module", "Title", "URL"])
    levels = ["A1-A2", "B1-B2", "C1-C2"]
    for level in levels:
        files = get_reading_lessons(base_dir, level)
        for f in files:
            title = f.replace('.js', '')
            url = f"https://ngoclong1209.github.io/ielts-reading-app/frontend/index.html?file={level}/{f}"
            ws.append(["Reading 1323", level, f"Reading {level} - {title}", url])
            
    # 2. Listening_Data
    sheet_name = "Listening_Data"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)
        
    ws.append(["Package", "Module", "Title", "URL"])
    listen_levels = ["Basic", "Intermediate", "Advanced"]
    for level in listen_levels:
        for i in range(1, 35):
            title = f"Listening {level} - Lesson {i}"
            url = f"https://ngoclong1209.github.io/pte-listening-audios/{level}/Lesson_{i}/index.html"
            ws.append(["Listening 3 Levels", level, title, url])

    # 3. FullTest_Data
    sheet_name = "FullTest_Data"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)
        
    ws.append(["Package", "Module", "Title", "URL"])
    
    # Listening full tests
    list_folder = os.path.join(base_dir, "data", "practicepteonline", "listening")
    if os.path.exists(list_folder):
        tests = [d for d in os.listdir(list_folder) if d.startswith("Test_")]
        tests.sort(key=lambda x: int(re.search(r'Test_(\d+)', x).group(1)))
        for t in tests:
            url = f"https://ngoclong1209.github.io/practicepteonline/index.html?type=listening&test={t}"
            ws.append(["Full Tests", "Listening", f"Full Test Listening - {t}", url])
            
    # Reading full tests
    read_folder = os.path.join(base_dir, "data", "practicepteonline", "reading")
    if os.path.exists(read_folder):
        tests = [d for d in os.listdir(read_folder) if d.startswith("Test_")]
        tests.sort(key=lambda x: int(re.search(r'Test_(\d+)', x).group(1)))
        for t in tests:
            url = f"https://ngoclong1209.github.io/practicepteonline/index.html?type=reading&test={t}"
            ws.append(["Full Tests", "Reading", f"Full Test Reading - {t}", url])

    print("Saving...")
    wb.save(excel_path)
    print("Done! Course data injected successfully.")

if __name__ == "__main__":
    inject_data()
