import os
import re
import time
import datetime
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook

# Cấu hình
TARGET_DIR = "/Users/vungoclong/Library/CloudStorage/GoogleDrive-ngoclong137129@gmail.com/My Drive/IELTS_LISTENING"
TARGET_FILE = "UPGRADE YOUR ILETS SKILLS.xlsx"
FILE_PATH = os.path.join(TARGET_DIR, TARGET_FILE)
PROJECT_DIR = "/Users/vungoclong/Desktop/Antigravity/UPGRADE YOUR ILETS LISTENING"

last_processed_time = 0

def get_lessons(base_dir, level):
    folder = os.path.join(base_dir, "ReadingA1-C2", "frontend", "data", level)
    files = [f for f in os.listdir(folder) if f.endswith(".js") and f.startswith("lesson_")]
    files.sort(key=lambda x: int(re.search(r'lesson_(\d+)', x).group(1)))
    return files

def add_business_days(start_date, add_days):
    current_date = start_date
    while add_days > 0:
        current_date += datetime.timedelta(days=1)
        if current_date.weekday() < 5:
            add_days -= 1
    return current_date

def process_new_students():
    global last_processed_time
    
    # Đợi 3 giây để Excel nhả file (tránh lỗi file bị khóa)
    time.sleep(3)
    
    try:
        wb = load_workbook(FILE_PATH)
    except Exception as e:
        print(f"Lỗi khi mở file: {e}")
        return

    if "Students" not in wb.sheetnames:
        return

    ws_students = wb["Students"]
    changes_made = False

    # Lấy dữ liệu bài tập
    levels = {
        "A1-A2": {"files": get_lessons(PROJECT_DIR, "A1-A2"), "quota": 15},
        "B1-B2": {"files": get_lessons(PROJECT_DIR, "B1-B2"), "quota": 15},
        "C1-C2": {"files": get_lessons(PROJECT_DIR, "C1-C2"), "quota": 13}
    }

    # Đọc từng dòng ở sheet Students (Dòng 1 là Tiêu đề)
    for row_idx in range(2, ws_students.max_row + 1):
        student_name = ws_students.cell(row=row_idx, column=2).value
        start_date_val = ws_students.cell(row=row_idx, column=3).value
        status = ws_students.cell(row=row_idx, column=4).value

        if student_name and status != "Đã tạo":
            print(f"🚀 Phát hiện học viên mới: {student_name}")
            
            # Xử lý ngày bắt đầu
            if start_date_val:
                if isinstance(start_date_val, datetime.datetime):
                    current_date = start_date_val.date()
                else:
                    try:
                        current_date = datetime.datetime.strptime(str(start_date_val).strip(), "%Y-%m-%d").date()
                    except:
                        current_date = datetime.datetime.now().date()
            else:
                current_date = datetime.datetime.now().date()

            if current_date.weekday() >= 5:
                current_date = add_business_days(current_date, 1 if current_date.weekday() == 6 else 2)

            # Lên lịch học
            schedule = []
            for level, data in levels.items():
                files = data["files"]
                quota = data["quota"]
                chunk = []
                for file in files:
                    chunk.append(file)
                    if len(chunk) == quota:
                        schedule.append((current_date, level, list(chunk)))
                        chunk = []
                        current_date = add_business_days(current_date, 1)
                if chunk:
                    schedule.append((current_date, level, list(chunk)))
                    current_date = add_business_days(current_date, 1)

            # Ghi vào sheet cá nhân
            if student_name in wb.sheetnames:
                ws = wb[student_name]
            else:
                if 'Kế hoạch' in wb.sheetnames:
                    ws = wb.copy_worksheet(wb['Kế hoạch'])
                    ws.title = student_name
                else:
                    ws = wb.create_sheet(student_name)
                    ws.append(['Ngày Học', 'Tên Bài Tập', 'Link Truy Cập', 'Trạng Thái'])

            last_row = ws.max_row
            while last_row > 1 and ws.cell(row=last_row, column=1).value is None and ws.cell(row=last_row, column=2).value is None:
                last_row -= 1
            
            start_row = last_row + 1
            base_url = "https://ngoclong1209.github.io/ielts-reading-app/frontend/index.html"
            
            for i, (date, level, files) in enumerate(schedule):
                for f in files:
                    ws.cell(row=start_row, column=1).value = date.strftime("%Y-%m-%d")
                    ws.cell(row=start_row, column=2).value = f"Reading {level} - {f.replace('.js', '')}"
                    
                    cell_link = ws.cell(row=start_row, column=3)
                    cell_link.value = "🔗 Làm bài ngay"
                    cell_link.hyperlink = f"{base_url}?file={level}/{f}"
                    cell_link.style = "Hyperlink"
                    
                    ws.cell(row=start_row, column=4).value = "Chưa làm"
                    start_row += 1

            # Đánh dấu đã tạo
            ws_students.cell(row=row_idx, column=4).value = "Đã tạo"
            changes_made = True
            print(f"✅ Đã tạo lịch thành công cho {student_name}")

    if changes_made:
        print("Đang lưu lại file Excel...")
        try:
            wb.save(FILE_PATH)
            last_processed_time = time.time()
            print("Lưu thành công. Đã đồng bộ lên hệ thống.")
        except Exception as e:
            print(f"Không thể lưu file: {e}")

class ExcelHandler(FileSystemEventHandler):
    def on_modified(self, event):
        global last_processed_time
        # Chỉ theo dõi file đích
        if not event.is_directory and event.src_path == FILE_PATH:
            # Ngăn chặn vòng lặp vô hạn do chính script save file
            if time.time() - last_processed_time > 10:
                print(f"🔄 File {TARGET_FILE} bị thay đổi. Đang kiểm tra...")
                process_new_students()

if __name__ == "__main__":
    print(f"👀 Đang chạy ngầm và theo dõi file: {FILE_PATH}...")
    event_handler = ExcelHandler()
    observer = Observer()
    observer.schedule(event_handler, TARGET_DIR, recursive=False)
    observer.start()
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()
