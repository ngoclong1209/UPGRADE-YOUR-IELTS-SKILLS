import openpyxl

EXCEL_FILE = 'UPGRADE YOUR ILETS SKILLS.xlsx'
wb = openpyxl.load_workbook(EXCEL_FILE)

# Cập nhật URLs trong Reading_Data, Listening_Data, FullTest_Data
for sheet_name in ['Reading_Data', 'Listening_Data', 'FullTest_Data']:
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Tìm cột URL
        url_col = None
        for col_idx, cell in enumerate(sheet[1], 1):
            if cell.value == 'URL':
                url_col = col_idx
                break
        
        if url_col:
            for row in range(2, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=url_col)
                if cell.value and isinstance(cell.value, str):
                    # Thay thế base url cũ
                    if 'https://ngoclong1209.github.io/' in cell.value and 'UPGRADE-YOUR-IELTS-SKILLS' not in cell.value:
                        cell.value = cell.value.replace('https://ngoclong1209.github.io/', 'https://ngoclong1209.github.io/UPGRADE-YOUR-IELTS-SKILLS/')
                    elif 'https://practicepteonline.github.io/' in cell.value:
                        # FullTest có thể mang link cũ practicepteonline.github.io
                        cell.value = cell.value.replace('https://practicepteonline.github.io/', 'https://ngoclong1209.github.io/UPGRADE-YOUR-IELTS-SKILLS/')
        print(f"Updated {sheet_name}")

wb.save(EXCEL_FILE)
print("Saved Excel file.")
